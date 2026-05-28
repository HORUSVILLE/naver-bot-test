"""
네이버 신규오픈 매장 자동 수집 봇 (GitHub Actions 호환 + 디버그 강화 버전)
- 네이버 지도(map.naver.com) 검색 → '새로오픈' 필터 적용
- 필터 적용을 3단계로 시도(칩 직접 클릭 → 패널 토글 → JS 강제 클릭)
- 실패 시 화면의 모든 버튼 텍스트와 HTML을 output에 저장(다음 진단용)
- '새로오픈' 태그가 있는 항목만 최종 저장
- 어떤 상황에서도 output 폴더 / summary.txt 는 반드시 생성됨
"""
from pathlib import Path
import os
import re
import sys
import time
import urllib.parse
import traceback

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from playwright.sync_api import sync_playwright


# ============================================================
# 검색어 — 추가/변경은 여기만 수정
# ============================================================
QUERIES = [
    "용인 맛집",
    "용인 카페",
    "용인 디저트",
    "용인 미용실",
    "용인 네일샵",
    "용인 펜션",
]

# ============================================================
# 동작 설정
# ============================================================
HEADLESS = True                # GitHub Actions에서는 반드시 True

MAX_PAGES_PER_QUERY = 10
PAGE_LOAD_WAIT_MS = 6000        # 페이지 진입 후 대기 (넉넉히)
FILTER_WAIT_MS = 4500
PAGE_CLICK_WAIT_MS = 3500
BETWEEN_QUERY_SEC = 7
SCROLL_MAX_STEPS = 80
SCROLL_STABLE_LIMIT = 4
ONLY_KEEP_NEW_OPEN = True

# 첫 검색어 실패 시 디버그 덤프를 남길지 (HTML + 버튼 텍스트)
DUMP_DEBUG_ON_FAIL = True

OUTPUT_DIR = Path("output")
SCREENSHOT_DIR = OUTPUT_DIR / "screenshots"
DEBUG_DIR = OUTPUT_DIR / "debug"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
DEBUG_DIR.mkdir(parents=True, exist_ok=True)
(OUTPUT_DIR / "summary.txt").write_text("초기화됨 — 아직 실행 전\n", encoding="utf-8")


# ============================================================
# 공통 유틸
# ============================================================
def safe_name(text: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]+", "_", text or "").strip("_")[:40]


def save_shot(page, filename: str):
    try:
        page.screenshot(path=str(SCREENSHOT_DIR / filename), full_page=True)
    except Exception:
        pass


def save_text(filename: str, content: str):
    try:
        (OUTPUT_DIR / filename).write_text(content, encoding="utf-8")
    except Exception:
        pass


def save_debug(filename: str, content: str):
    try:
        (DEBUG_DIR / filename).write_text(content, encoding="utf-8")
    except Exception:
        pass


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def split_lines(text: str):
    if not text:
        return []
    return [normalize(x) for x in re.split(r"[\r\n]+", text) if normalize(x)]


# ============================================================
# 진단 덤프 — 화면의 클릭 가능한 요소 텍스트 전부 기록
# ============================================================
def dump_clickable_texts(page, query, tag):
    """페이지의 button/a/[role=button] 텍스트를 모아 파일로 저장."""
    try:
        texts = page.evaluate(
            """() => {
                const out = [];
                const els = document.querySelectorAll("button, a, [role='button'], span");
                for (const el of els) {
                    const t = (el.innerText || el.textContent || '').trim();
                    if (t && t.length <= 30) out.push(t);
                }
                return Array.from(new Set(out));
            }"""
        )
        save_debug(
            f"clickable_{safe_name(query)}_{tag}.txt",
            "\n".join(texts),
        )
    except Exception as e:
        save_debug(f"clickable_{safe_name(query)}_{tag}_ERROR.txt", str(e))


def dump_html(page, query, tag):
    try:
        html = page.content()[:200000]
        save_debug(f"html_{safe_name(query)}_{tag}.html", html)
    except Exception:
        pass


# ============================================================
# 검색결과 iframe(frame) 찾기
# ============================================================
def get_search_frame(page, timeout_sec=30):
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        for fr in page.frames:
            url = fr.url or ""
            if "pcmap.place.naver.com" in url and "/list" in url:
                return fr
        page.wait_for_timeout(500)
    raise RuntimeError("검색결과 iframe을 찾지 못함")


# ============================================================
# '새로오픈' 필터 적용 (3단계 시도)
# ============================================================
def try_click_text(scope, text, exact=False, timeout=3500):
    """scope(page 또는 frame) 안에서 text를 가진 클릭요소를 시도."""
    patterns = [
        f"button:has-text('{text}')",
        f"a:has-text('{text}')",
        f"[role='button']:has-text('{text}')",
        f"label:has-text('{text}')",
        f"span:has-text('{text}')",
    ]
    for sel in patterns:
        try:
            loc = scope.locator(sel)
            n = loc.count()
            for i in range(min(n, 5)):
                el = loc.nth(i)
                if exact:
                    t = (el.inner_text(timeout=1000) or "").strip()
                    if t != text:
                        continue
                el.scroll_into_view_if_needed(timeout=2000)
                el.click(timeout=timeout)
                return True
        except Exception:
            continue
    return False


def js_force_click(page, text):
    """JS로 해당 텍스트를 가진 요소를 강제 클릭 (CSS 셀렉터가 안 먹을 때)."""
    try:
        return page.evaluate(
            """(label) => {
                const els = Array.from(document.querySelectorAll("button, a, [role='button'], span, label"));
                for (const el of els) {
                    const t = (el.innerText || el.textContent || '').trim();
                    if (t === label || t.startsWith(label)) {
                        el.click();
                        return true;
                    }
                }
                return false;
            }""",
            text,
        )
    except Exception:
        return False


def apply_new_open_filter(page, query):
    """
    1단계: '새로오픈' 칩/버튼을 페이지에서 바로 클릭
    2단계: 필터 패널을 열고 → 패널 내부를 스크롤하며 '새로오픈' 토글 → '결과보기'
    3단계: JS 강제 클릭
    """
    page.wait_for_timeout(1500)

    # ---- 1단계: 바로 클릭 ----
    if try_click_text(page, "새로오픈"):
        page.wait_for_timeout(FILTER_WAIT_MS)
        print(f"[{query}] 1단계: '새로오픈' 직접 클릭 성공")
        return True

    # ---- 2단계: 필터 패널 경로 ----
    print(f"[{query}] 1단계 실패 → 2단계(필터 패널) 시도")

    # 필터 패널 열기: '필터' 텍스트 or aria-label
    opened = (
        try_click_text(page, "필터")
        or _click_filter_icon(page)
    )
    page.wait_for_timeout(1500)

    if opened:
        # 패널 내부를 단계적으로 스크롤하며 '새로오픈' 탐색
        for step in range(8):
            if try_click_text(page, "새로오픈"):
                page.wait_for_timeout(800)
                # 결과보기 클릭
                if not (try_click_text(page, "결과보기") or try_click_text(page, "적용")):
                    js_force_click(page, "결과보기")
                page.wait_for_timeout(FILTER_WAIT_MS)
                print(f"[{query}] 2단계: 패널에서 '새로오픈' 적용 성공 (scroll step={step})")
                return True
            # 패널 영역으로 추정되는 곳에 마우스를 올리고 스크롤
            try:
                page.mouse.move(640, 600)
                page.mouse.wheel(0, 350)
                page.wait_for_timeout(500)
            except Exception:
                break

    # ---- 3단계: JS 강제 클릭 ----
    print(f"[{query}] 2단계 실패 → 3단계(JS 강제 클릭) 시도")
    if js_force_click(page, "새로오픈"):
        page.wait_for_timeout(1200)
        if not js_force_click(page, "결과보기"):
            try_click_text(page, "결과보기")
        page.wait_for_timeout(FILTER_WAIT_MS)
        print(f"[{query}] 3단계: JS 강제 클릭 성공")
        return True

    # ---- 전부 실패 → 진단 정보 덤프 ----
    if DUMP_DEBUG_ON_FAIL:
        dump_clickable_texts(page, query, "fail")
        dump_html(page, query, "fail")
        save_shot(page, f"{safe_name(query)}_FAIL_filter.png")

    raise RuntimeError("'새로오픈' 필터를 적용하지 못함 (3단계 모두 실패)")


def _click_filter_icon(page):
    """슬라이더 모양 필터 아이콘을 여러 방법으로 클릭 시도."""
    selectors = [
        "button[aria-label*='필터']",
        "button[class*='filter']",
        "a[class*='filter']",
        "button[class*='Filter']",
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0:
                loc.click(timeout=3000)
                return True
        except Exception:
            continue
    return False


def open_search_with_new_open(page, query: str):
    encoded = urllib.parse.quote(query)
    map_url = f"https://map.naver.com/p/search/{encoded}?searchType=place"

    page.goto(map_url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(PAGE_LOAD_WAIT_MS)
    save_shot(page, f"{safe_name(query)}_01_map.png")

    # 첫 검색어는 진단을 위해 클릭가능 요소 텍스트를 항상 덤프
    if DUMP_DEBUG_ON_FAIL and query == QUERIES[0]:
        dump_clickable_texts(page, query, "initial")

    apply_new_open_filter(page, query)
    save_shot(page, f"{safe_name(query)}_02_filtered.png")

    frame = get_search_frame(page, timeout_sec=30)
    return frame


# ============================================================
# 카드 수집
# ============================================================
SCROLL_SELECTORS = [
    "#_pcmap_list_scroll_container",
    "div[id*='_list_scroll_container']",
    "div.Ryr1F",
]

CARD_SELECTORS = [
    "ul > li:has(a)",
    "li:has(a):has(span)",
    "li.UEzoS",
    "li.VLTHu",
]

LABELS = {
    "새로오픈", "예약", "주문", "톡톡", "쿠폰", "포장", "광고",
    "영업 종료", "영업종료", "영업중", "영업 중", "place+", "Npay+", "Npay",
    "단체석", "주차", "리뷰많은", "요즘뜨는", "신선한", "분위기좋은",
}

CATEGORY_HINTS = {
    "카페", "디저트", "음식점", "한식", "중식", "양식", "일식", "분식",
    "이자카야", "베이커리", "피자", "치킨", "고기", "국밥", "라면",
    "미용실", "네일", "마사지", "에스테틱", "헤어", "바버샵",
    "펜션", "숙박", "호텔", "모텔", "게스트하우스", "리조트", "풀빌라",
    "주꾸미요리", "백숙", "삼계탕", "중식당",
}


def find_first_existing(ctx, selectors):
    for sel in selectors:
        try:
            if ctx.locator(sel).count() > 0:
                return sel
        except Exception:
            continue
    return None


def is_address_line(line: str) -> bool:
    if re.search(r"(시|군|구|동|읍|면|로|길|가)\s*[\d가-힣]", line):
        return True
    if re.search(r"\d+[-\d]*번지", line):
        return True
    if re.match(r"^[가-힣\s]+(시|군)\s+[가-힣]+(구|동|읍|면)", line):
        return True
    return False


def is_review_line(line: str) -> bool:
    return "리뷰" in line and re.search(r"\d", line) is not None


def extract_card_info(card, query, page_num, rank):
    try:
        text = card.inner_text(timeout=2000)
    except Exception:
        return None

    lines = split_lines(text)
    if len(lines) < 2:
        return None
    if all(line in LABELS for line in lines):
        return None

    name = lines[0]
    if name in LABELS or len(name) < 2:
        return None

    is_new_open = any("새로오픈" in ln for ln in lines)

    category = ""
    address = ""
    review = ""
    extras = []

    for line in lines[1:]:
        if line in LABELS:
            continue
        if not address and is_address_line(line):
            address = line
            continue
        if not review and is_review_line(line):
            review = line
            continue
        if not category and any(h in line for h in CATEGORY_HINTS) and len(line) <= 25:
            category = line
            continue
        extras.append(line)

    href = ""
    try:
        href = card.locator("a").first.get_attribute("href") or ""
    except Exception:
        pass

    return {
        "query": query,
        "page": page_num,
        "rank": rank,
        "name": name,
        "category": category,
        "address": address,
        "review": review,
        "new_open": "Y" if is_new_open else "",
        "extras": " | ".join(extras[:5]),
        "href": href,
        "raw": " / ".join(lines),
    }


def collect_cards(frame, query, page_num):
    card_sel = find_first_existing(frame, CARD_SELECTORS)
    if not card_sel:
        return []

    cards = frame.locator(card_sel)
    count = cards.count()

    rows = []
    rank = 0
    for i in range(count):
        info = extract_card_info(cards.nth(i), query, page_num, rank + 1)
        if info:
            rank += 1
            rows.append(info)
    return rows


def scroll_list_to_end(page, frame):
    scroller = find_first_existing(frame, SCROLL_SELECTORS) or "body"
    last_h = -1
    same = 0

    for _ in range(SCROLL_MAX_STEPS):
        try:
            frame.evaluate(
                "(sel) => { const el = sel === 'body' ? document.scrollingElement : document.querySelector(sel); if (el) el.scrollTop = el.scrollHeight; }",
                scroller,
            )
        except Exception:
            break

        page.wait_for_timeout(900)

        try:
            cur_h = frame.evaluate(
                "(sel) => { const el = sel === 'body' ? document.scrollingElement : document.querySelector(sel); return el ? el.scrollHeight : 0; }",
                scroller,
            )
        except Exception:
            break

        if cur_h == last_h:
            same += 1
        else:
            same = 0
            last_h = cur_h
        if same >= SCROLL_STABLE_LIMIT:
            break


# ============================================================
# 페이지 순회
# ============================================================
def click_page_number(frame, page_num):
    selectors = [
        f"a:has-text('{page_num}')",
        f"button:has-text('{page_num}')",
        f"span:has-text('{page_num}')",
    ]
    for sel in selectors:
        try:
            loc = frame.locator(sel)
            n = loc.count()
            for i in range(n):
                el = loc.nth(i)
                txt = (el.inner_text(timeout=1000) or "").strip()
                if txt == str(page_num):
                    el.scroll_into_view_if_needed(timeout=2000)
                    el.click(timeout=3000)
                    return True
        except Exception:
            continue
    return False


def collect_all_pages(page, frame, query):
    all_rows = []
    seen = set()

    scroll_list_to_end(page, frame)
    save_shot(page, f"{safe_name(query)}_p1_end.png")
    rows = collect_cards(frame, query, 1)
    for r in rows:
        key = (r["query"], r["name"], r["address"])
        if key in seen:
            continue
        seen.add(key)
        all_rows.append(r)
    print(f"[{query}] p1 수집: {len(rows)}건")

    for p in range(2, MAX_PAGES_PER_QUERY + 1):
        if not click_page_number(frame, p):
            print(f"[{query}] p{p} 버튼 없음 — 종료")
            break

        page.wait_for_timeout(PAGE_CLICK_WAIT_MS)
        try:
            frame = get_search_frame(page, timeout_sec=15)
        except Exception:
            pass

        scroll_list_to_end(page, frame)
        save_shot(page, f"{safe_name(query)}_p{p}_end.png")

        rows = collect_cards(frame, query, p)
        if not rows:
            print(f"[{query}] p{p} 카드 없음 — 종료")
            break

        added = 0
        for r in rows:
            key = (r["query"], r["name"], r["address"])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(r)
            added += 1

        print(f"[{query}] p{p} 추가: {added}건 (누적 {len(all_rows)})")
        if added == 0:
            break

    return all_rows


def run_one_query(page, query):
    print(f"\n{'='*50}\n[START] {query}\n{'='*50}")
    frame = open_search_with_new_open(page, query)
    rows = collect_all_pages(page, frame, query)

    if ONLY_KEEP_NEW_OPEN:
        before = len(rows)
        rows = [r for r in rows if r["new_open"] == "Y"]
        print(f"[{query}] 새로오픈 필터링: {before}건 → {len(rows)}건")

    print(f"[DONE] {query}: 총 {len(rows)}건")
    return rows


# ============================================================
# 엑셀 저장
# ============================================================
def save_excel(all_rows, summary_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["검색어", "페이지", "순위", "상호명", "카테고리", "주소",
               "리뷰", "새로오픈", "참고사항", "링크", "원본텍스트"]
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="305496")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in all_rows:
        ws.append([
            r["query"], r["page"], r["rank"],
            r["name"], r["category"], r["address"],
            r["review"], r["new_open"], r["extras"], r["href"], r["raw"],
        ])

    widths = [12, 6, 6, 30, 18, 40, 14, 8, 35, 50, 60]
    for i, w in enumerate(widths, 1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2.append(["항목", "값"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
        cell.fill = header_fill
    for k, v in summary_data:
        ws2.append([k, str(v)])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 70

    wb.save(OUTPUT_DIR / "results.xlsx")


# ============================================================
# main
# ============================================================
def main():
    all_rows = []
    summary = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--lang=ko-KR",
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
            ],
        )
        context = browser.new_context(
            viewport={"width": 1600, "height": 1200},
            locale="ko-KR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            ),
        )

        for idx, query in enumerate(QUERIES):
            page = context.new_page()
            try:
                rows = run_one_query(page, query)
                all_rows.extend(rows)
                summary.append((f"[{query}] 상태", "성공"))
                summary.append((f"[{query}] 수집건수", len(rows)))
            except Exception as e:
                save_shot(page, f"ERROR_{safe_name(query)}.png")
                save_text(f"error_{safe_name(query)}.txt", traceback.format_exc())
                summary.append((f"[{query}] 상태", f"실패: {e}"))
                print(traceback.format_exc())
            finally:
                page.close()

            if idx < len(QUERIES) - 1:
                time.sleep(BETWEEN_QUERY_SEC)

        browser.close()

    summary.append(("== 총 수집 건수 ==", len(all_rows)))
    save_excel(all_rows, summary)

    summary_lines = [f"총 수집: {len(all_rows)}건", "", "[검색어별 결과]"]
    for k, v in summary:
        summary_lines.append(f"  {k}: {v}")
    save_text("summary.txt", "\n".join(summary_lines))

    print(f"\n[FINAL] 총 {len(all_rows)}건 수집 완료")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        err = traceback.format_exc()
        print(err)
        save_text("summary.txt", f"치명적 오류로 중단됨:\n\n{err}")
        save_text("fatal_error.txt", err)
        sys.exit(0)
